"""EMA EPAR (European Public Assessment Report) client.

EMA doesn't have a clean REST API like openFDA. They publish a bulk medicines
list as Excel/CSV at https://www.ema.europa.eu/en/medicines/download-medicine-data.
We download it once per refresh, cache to memory, and filter per drug query.

Output: list of EMAApproval records — one per matching authorized medicine.

Trade-off: this is a single bulk download per refresh (one HTTP request total
across all antigens) instead of per-antigen API calls. Caching the parsed table
across antigens makes this very cheap.
"""

import asyncio
import csv
import io
from datetime import date, datetime, timezone
from typing import Any

import httpx

from taa.schema import Antigen, Citation, EMAApproval, SourceFreshness

# EMA's bulk "Medicines: all" report — XLSX with all human + veterinary medicines.
# URL discovered by scraping https://www.ema.europa.eu/en/medicines/download-medicine-data
# (paths change periodically; if 404 again, re-scrape that page for current path).
EMA_DIRECT_CSV = "https://www.ema.europa.eu/en/documents/report/medicines-output-medicines-report_en.xlsx"

TIMEOUT_S = 60
RETRY_ATTEMPTS = 3

_lock = asyncio.Lock()
_cache: dict[str, Any] = {"loaded": False, "rows": []}


class EMAResult:
    def __init__(
        self,
        approvals: list[EMAApproval],
        citations: list[Citation],
        freshness: SourceFreshness,
    ) -> None:
        self.approvals = approvals
        self.citations = citations
        self.freshness = freshness


async def fetch_for_drugs(
    antigen: Antigen,
    drug_names: list[str],
    citation_id_start: int = 1,
) -> EMAResult:
    """Look up EMA EPARs for known drug names. Caches the EMA medicines list."""
    attempt_at = datetime.now(timezone.utc)

    if not drug_names:
        return EMAResult(
            approvals=[],
            citations=[],
            freshness=SourceFreshness(
                source="ema", last_attempt=attempt_at, last_success=attempt_at
            ),
        )

    try:
        rows = await _load_ema_table()
    except (httpx.HTTPError, httpx.TimeoutException, ValueError) as e:
        return EMAResult(
            approvals=[],
            citations=[],
            freshness=SourceFreshness(
                source="ema", last_attempt=attempt_at, error=f"{type(e).__name__}: {e}"
            ),
        )

    # Word-boundary match against name/INN/active substance to avoid generic
    # markers like "CAR-T" or "vaccine" exploding into hundreds of EMA hits.
    # Skip very-short or generic markers that would over-match.
    GENERIC_MARKERS = {
        "car-t", "car t", "chimeric antigen receptor", "vaccine",
        "peptide vaccine", "dendritic cell", "radioligand",
    }
    safe_names = [
        d for d in drug_names
        if len(d) >= 4 and d.lower() not in GENERIC_MARKERS
    ]
    if not safe_names:
        return EMAResult(
            approvals=[],
            citations=[],
            freshness=SourceFreshness(
                source="ema", last_attempt=attempt_at, last_success=attempt_at
            ),
        )

    import re as _re
    patterns = [_re.compile(rf"\b{_re.escape(d)}\b", _re.IGNORECASE) for d in safe_names]
    matched_rows: list[dict[str, str]] = []
    seen_keys: set[str] = set()
    for row in rows:
        haystack = (
            (row.get("name") or "")
            + " "
            + (row.get("inn") or "")
            + " "
            + (row.get("active_substance") or "")
        )
        if any(p.search(haystack) for p in patterns):
            key = row.get("ema_product_number", "") or row.get("name", "")
            if key in seen_keys:
                continue
            seen_keys.add(key)
            matched_rows.append(row)

    approvals = [_normalize(r) for r in matched_rows]
    approvals = [a for a in approvals if a is not None]
    citations = _build_citations(approvals, citation_id_start, attempt_at)  # type: ignore[arg-type]

    return EMAResult(
        approvals=approvals,  # type: ignore[arg-type]
        citations=citations,
        freshness=SourceFreshness(
            source="ema", last_success=attempt_at, last_attempt=attempt_at
        ),
    )


async def _load_ema_table() -> list[dict[str, str]]:
    """Download + parse the EMA medicines table once per process."""
    async with _lock:
        if _cache["loaded"]:
            return _cache["rows"]  # type: ignore[no-any-return]

        async with httpx.AsyncClient(
            timeout=TIMEOUT_S,
            headers={"User-Agent": "TAA Tracker (rsingla92@gmail.com)"},
            follow_redirects=True,
        ) as client:
            for attempt in range(RETRY_ATTEMPTS):
                try:
                    resp = await client.get(EMA_DIRECT_CSV)
                    resp.raise_for_status()
                    rows = _parse_ema_xlsx(resp.content)
                    _cache["loaded"] = True
                    _cache["rows"] = rows
                    return rows
                except httpx.HTTPError:
                    if attempt == RETRY_ATTEMPTS - 1:
                        raise
                    await asyncio.sleep(2.0)

    return []


def _parse_ema_xlsx(content: bytes) -> list[dict[str, str]]:
    """Parse the EMA medicines XLSX file.

    EMA's spreadsheet has a header row at line 9 (rows 1-8 are metadata).
    We use openpyxl in read-only streaming mode to keep memory low.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        # Fall back: try to parse as CSV/TSV in case the URL changed
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        return _parse_ema_csv(text)

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows_iter = ws.iter_rows(values_only=True)
    # Skip header rows until we find one starting with "Medicine name" or "Name of medicine"
    header: list[str] | None = None
    data_rows: list[list[Any]] = []
    for row in rows_iter:
        if not row:
            continue
        first = str(row[0] or "").strip().lower()
        if header is None:
            if first.startswith(("medicine name", "name of medicine", "category")):
                header = [str(c or "").strip() for c in row]
            continue
        data_rows.append(list(row))

    if header is None:
        return []

    # Map EMA column names to our normalized keys. Names change between report
    # revisions; the mapping is lenient so it survives small renames.
    col_map = {}
    for i, col in enumerate(header):
        col_lc = col.lower()
        if "name of medicine" in col_lc or "medicine name" in col_lc:
            col_map["name"] = i
        elif "active substance" in col_lc:
            col_map["active_substance"] = i
        elif "international non-proprietary name" in col_lc:
            col_map["inn"] = i
        elif "ema product number" in col_lc or col_lc == "product number":
            col_map["ema_product_number"] = i
        elif "marketing authorisation date" in col_lc:
            col_map["authorisation_date"] = i
        elif "marketing authorisation" in col_lc and (
            "holder" in col_lc or "applicant" in col_lc or "developer" in col_lc
        ):
            col_map["mah"] = i
        elif "atc code" in col_lc:
            col_map["atc"] = i
        elif "medicine url" in col_lc or col_lc == "url":
            col_map["url"] = i

    rows: list[dict[str, str]] = []
    for r in data_rows:
        rows.append({key: str(r[i] or "").strip() for key, i in col_map.items() if i < len(r)})
    return rows


def _parse_ema_csv(text: str) -> list[dict[str, str]]:
    """Fallback CSV parser if openpyxl isn't available."""
    reader = csv.DictReader(io.StringIO(text), delimiter=";")
    return [{k.strip().lower().replace(" ", "_"): v for k, v in row.items() if k} for row in reader]


def _normalize(row: dict[str, str]) -> EMAApproval | None:
    name = row.get("name") or ""
    if not name:
        return None
    auth_date_str = row.get("authorisation_date", "")
    auth_date: date | None = None
    if auth_date_str:
        try:
            # EMA dates: "DD/MM/YYYY" or ISO
            if "/" in auth_date_str:
                d, m, y = auth_date_str.split("/")
                auth_date = date(int(y), int(m), int(d))
            else:
                auth_date = date.fromisoformat(auth_date_str[:10])
        except (ValueError, IndexError):
            pass

    return EMAApproval(
        name=name,
        active_substance=row.get("active_substance") or row.get("inn") or "",
        marketing_authorisation_holder=row.get("mah") or "",
        ema_product_number=row.get("ema_product_number") or "",
        authorisation_date=auth_date,
        atc_code=row.get("atc") or None,
        url=row.get("url") or f"https://www.ema.europa.eu/en/medicines/human/EPAR/{_slugify(name)}",
    )


def _slugify(name: str) -> str:
    return name.lower().replace(" ", "-").replace("/", "-")


def _build_citations(
    approvals: list[EMAApproval],
    citation_id_start: int,
    retrieved_at: datetime,
) -> list[Citation]:
    citations: list[Citation] = []
    cite_id = citation_id_start
    for app in approvals:
        url = app.url or f"https://www.ema.europa.eu/en/medicines/human/EPAR/{_slugify(app.name)}"
        citations.append(
            Citation(
                id=cite_id,
                url=url,  # type: ignore[arg-type]
                title=f"EMA EPAR · {app.name} · {app.marketing_authorisation_holder}",
                source_type="filing",
                retrieved_at=retrieved_at,
                locator=f"EMA {app.ema_product_number or app.name}",
            )
        )
        cite_id += 1
    return citations
